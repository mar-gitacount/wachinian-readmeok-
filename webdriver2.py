
# -*- coding: utf-8 -*-

from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
import sys
from selenium.webdriver.chrome.options import Options
num = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# ここで num を使用して何かを実行する
(f"受け取った引数: {num}")
url = "https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_Spec106=1,2&pdf_vi=c"
url = f"https://nanboya.com/search/item-list/b-948/?page={num}"
url = f"https://buy.watchnian.com/brand_rolex/"


# ?モデル、ブレスレット、文字盤を抽出する関数
def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    beltpattern = r"\[(.*?)\]|\((.*?)\)"
    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。
    items = {"model": model, "beltmatches": beltmatches}
    return items

def get_page_source(url):
    """
    Given a URL, use WebDriver to access the page, retrieve its source code,
    and return a BeautifulSoup object.
    """
    options = Options()
    options.add_argument("--headless")  
    
    # WebDriverのインスタンスを作成
    driver = webdriver.Chrome(options=options)  # Chromeを使用する場合。他のブラウザを使う場合は適宜変更してください。

    # 指定されたURLにアクセス
    driver.get(url)

    # options = Options()
    # options.add_argument("--headless")  # ヘッドレスモードを有効にする
    # driver = webdriver.Chrome(options=options)  # または他のブラウザに合わせて選択




    # ページのソースコードを取得
    page_source = driver.page_source

    # ページのソースコードをBeautiful Soupオブジェクトに変換
    soup = BeautifulSoup(page_source, "html.parser")

    # WebDriverを終了
    driver.quit()

    return soup

# htmlを返す関数
def parse_html(html_string):
    """
    Given an HTML string, parse it and return a BeautifulSoup object.
    """
    return BeautifulSoup(html_string, "html.parser")




# ?エクセルに入力する関数
def wsinsert(values, sheet):
    print("wsinsert関数", values)
    sheet.append(values)
    # for item in values:
    #     sheet.append(item)


# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
# ファイル名に日付を組み込む

file_name = f"ウォッチニアン_{today_date}買取.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(
        [
            "モデル名",
            "リファレンスNO",
            "文字盤",
            "買取価格",
            "URL",
        ]
    )
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active


# SeleniumのWebDriverを初期化
options = Options()
options.add_argument("--headless")  # ヘッドレスモードを有効にする
driver = webdriver.Chrome(options=options)  # !または他のブラウザに合わせて選択。本番はこれを実行して、バックグラウンド通信する。
driver = webdriver.Chrome()  # または他のブラウザに合わせて選択

# URLを開く
driver.get(url)

# Seleniumがページのロードを待つなどの適切な待機処理が必要な場合はここで実施

# ページのHTMLを取得
page_source = driver.page_source

# BeautifulSoupを使ってHTMLを解析
soup = BeautifulSoup(page_source, "html.parser")


# !ここから処理スタート
#?入稿可能な変数
# <tbody> タグ内のテキストを抽出して表示
tbody_tag = soup.find("body")
# 各モデルページを取得する。
models_get = tbody_tag.find_all("div",class_="editorTmplBnrs_item")

# print(models_get)
# ?各モデル一覧。

# ?以下リファレンスナンバー正規表現
ref_pattern = r"\b(\d{3,}[A-Z]*)\b"
# ?アルファベットを抽出する正規表現
lot_pattern =r"\b[A-Za-z]+\b"
lot_pattern =r"\b[A-Za-z]|[A-Za-z]/+[A-Za-z]+\b"
# ?日本語を抽出するパターン。金額を抽出する際に利用する。
japanese_pattern = r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF]+' 
color_pattern = r"\b(ダークロジウム|Zブルー|チョコ|シャンパン・黒文字盤|スチール・黒文字盤|青文字盤|シャンパン文字盤|ブラック文字盤|ブラックシェル文字盤|シャンパン黒目|サンダスト文字盤|ホワイトアラビア文字盤|白文字盤|黒文字盤|アイスブルー文字盤|ラピスラズリグリーンパーム|チョコレートコンピューター|オーベルジーヌ VI IXダイヤ|ゴールデンフルーテッドモチーフ|ソーダライト|ロゼ|スパイダーダイアル|オールトリチウム|トリチノバ|オリーブグリーン|パヴェダイヤ|シャンパンフルーテッド|ブラックシェル|NEWダイヤル|ロゼローマ |スチール|ブルーローマ|新ブルーダイヤル|ホワイトシェル|シルバー スターダイヤ |チョコレート|ムーンフェイズ|スレート|Dブルー|アイボリー文字盤|ブライトブルー|ターコイズブルーセレブレーション モチーフ|オイスター|キャンディピンク|ホワイトローマ|ダークロジウム|ターコイズブルーセレブレーション|ターコイズブルー|セレブレーションモチーフ|ブラック コンセントリック|コーラルレッド|レッド|ブルー|ブラック|イエロー|グリーン|グレー|黒|ダークグレー|ホワイト|シャンパン|シルバー|ゴールド|アイスブルー|ブラウン|スレートローマ|ブルーブラック)*\b"

ref_pattern = r"\b\d{3,}[A-Z]*\b"
for  model in models_get:
    # print(model)
    a_tag_get = model.find("a").get("href")
    useurl = "https://buy.watchnian.com"
    # 各モデルのurl
    model_url = useurl + a_tag_get
    # print(model_url)
    modelpage = get_page_source(model_url)
    refitems = modelpage.find("ul",class_="brandList03_list")
    refitem = refitems.find_all("li",class_="brandList03_item")
    refitems = modelpage.find_all("li",class_="brandList03_item")

    for item in refitems:
        # ?リファレンスナンバー内ループ、URLを取得する。
        r = item.find("a").get("href")
        refitemurl = useurl + r
        # print(r)
        getitemspagesoup = get_page_source(refitemurl)
        # print(getitemspagesoup)
        # ?以下でエラーの場合、買取が存在しない。
        try:
            getitemspage_body_tag = getitemspagesoup.find("body").find_all("div","casestudyList04_block")
        except Exception as e:
            continue
        # brandList03_item_href_dateil_div_tag = getitemspage_body_tag.find("body").find_all("div","casestudyList04_block")
        for wathcItem in getitemspage_body_tag:
             #  ?以下存在しない場合は、買取なし
             try:
                 maindetaill = wathcItem.find("p",class_="casestudyList04_title02").getText(strip=True)
                 # ?モデル名
                 maindetaillWords = maindetaill.split()
                 modelname = maindetaillWords[0]
                 print(maindetaill)
             except Exception as e:
                 continue
             try:
                 newprice = wathcItem.find("dl",class_="casestudyList04_conflictSet-brandNew").find("dd").getText(strip=True)
                 newprice = re.sub(japanese_pattern,"",newprice)
                #  newprice = int(newprice)
                 print(newprice)
             except Exception as e:
                 newprice = ""
             try:
                   usedprice = wathcItem.find("dl",class_="casestudyList04_conflictSet-used").getText(strip=True)
                   usedprice = re.sub(japanese_pattern,"",usedprice)
                   print(usedprice)
             except Exception as e:
                 usedprice = ""
             ref_get = re.findall(ref_pattern,maindetaill)
             code_get = re.findall(lot_pattern,maindetaill)
             dispray = re.findall(color_pattern,maindetaill)
            #  dispray = re.sub(japanese_pattern,"",dispray)
            # リファレンス取得
            #ref_join = ''.join(ref_get)
            #print(ref_join)
             
            # ディスプレイ取得
             disprayjoin = ''.join(dispray)
             if len(ref_get) > 2:
                ref_join = ref_get[1]
             else:
                ref_join = ref_get[0]
            
             print(code_get)
             insertitems=[modelname,ref_join,disprayjoin,newprice,usedprice,refitemurl]
             wsinsert(insertitems,ws)
        #         print(maindetaillWords)
        #         print(code_get)
        #         print(dispray)
        #         print("--------------------------")
             wb.save(file_name)

            #dispray = re.findall(color_pattern,maindetaill)
             print("-------------------------------------")
        # brandList03_itemInner = getitemspagesoup.find("a",class_="brandList03_itemInner")
        # # ?リファレンスナンバーごとにURLを取得する
        # another_refitems = useurl + brandList03_itemInner.get("href")
        # another_refitems_pages = get_page_source(another_refitems)
        # # print(another_refitems_pages)
        # # ?リファレンス内のアイテム一覧
        # # ? なぜかアイテム数分だけ余計にループする
        # brandList03_item_href_dateil_div_tag = another_refitems_pages.find("body").find_all("div","casestudyList04_block")
        # print(len(brandList03_item_href_dateil_div_tag))
        # for wathcItem in brandList03_item_href_dateil_div_tag:
        #      newprice = wathcItem.find("dl",class_="casestudyList04_conflictSet-brandNew").find("dd").getText(strip=True)
        #      print(newprice)
        #      print(refitemurl)
        #      print(another_refitems)
        # print(brandList03_item_href_dateil_div_tag)
        # try:
        #     newprice =  brandList03_item_href_dateil_div_tag.find("dl",class_="casestudyList04_conflictSet-brandNew").find("dd").getText(strip=True)
        #     print(newprice)
        # except Exception as e:
        #             newprice = ""

        # for brandList03_item in brandList03_itemInner:
        #     brandList03_item_href = useurl +  brandList03_item.get("href")
        #     # print(brandList03_item_href)
        #     brandList03_item_href_dateil = get_page_source(brandList03_item_href)
        #     # 各時計に到達
        #     # print(brandList03_item_href_dateil)
        #     try:
        #         brandList03_item_href_dateil_div_tag = brandList03_item_href_dateil.find("body").find_all("div","casestudyList04_block")
        #     except Exception as e:
        #         print("ここでエラーになる場合は、アイテムが存在しない。")
        #         continue

        #     for brandList03_item_href_dateil_div_tag_item in brandList03_item_href_dateil_div_tag:
        #         # print(brandList03_item_href_dateil_div_tag_item)
        #         # dateillpage = get_page_source(brandList03_item_href_dateil_div_tag_item)
        #         maindetaill = brandList03_item_href_dateil_div_tag_item.find("p",class_="casestudyList04_title02").getText(strip=True)
        #         try:
        #             newprice =  brandList03_item_href_dateil_div_tag_item.find("dl",class_="casestudyList04_conflictSet-brandNew").find("dd").getText(strip=True)
        #             print(newprice)
        #         except Exception as e:
        #             newprice = ""
               
        #         try:
        #             usedprice = brandList03_item_href_dateil_div_tag_item.find("dl",class_="casestudyList04_conflictSet-used").getText(strip=True)
                
        #         except Exception as e:
        #             usedprice = ""
        #         # 
        #         ref_get = re.findall(ref_pattern,maindetaill)
        #         code_get = re.findall(lot_pattern,maindetaill)
        #         dispray = re.findall(color_pattern,maindetaill)
               
        #         # リファレンス取得
        #         ref_join = ''.join(ref_get)   
        #         print(ref_join)
        #         # ディスプレイ取得
        #         disprayjoin = ''.join(dispray)

        #         non_empty_values = [value for value in dispray if value.strip() != '']
        #         print(disprayjoin)
        #         # モデル名
        #         maindetaillWords = maindetaill.split()
        #         modelname = maindetaillWords[0]
        #         blesret_etc = maindetaill.replace(modelname,"")
        #         blesret_etc = blesret_etc.replace(disprayjoin,"")
        #         blesret_etc = blesret_etc.replace(ref_join,"")
                
        #         #URL    
        #         brandList03_item_href
        #         print(maindetaill)
        #         insertitems=[modelname,ref_join,disprayjoin,newprice,usedprice,brandList03_item_href]
        #         wsinsert(insertitems,ws)
        #         print(maindetaillWords)
        #         print(code_get)
        #         print(dispray)
        #         print("--------------------------")
        #         wb.save(file_name)

        # conda activate python_env3.9 cd Desktop cd 仕事でつかうやつ cd 買取 cd ウォッチニアン python subtest.py
    # driver.get(model_url)
    # brandList03_list

  








